"""
Demo/sample workbooks for Excel AI Analyzer — designed to IMPRESS a trial user
and show what the local parser does that a naive tool can't. Fake data by
construction (safe to ship). Realistic finance layouts.

Each file has a matching set of 'try asking...' questions (printed) so the
landing page / first-run can suggest them.
"""
import random, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
random.seed(7)
D="files"; os.makedirs(D, exist_ok=True)

NAVY=PatternFill("solid", fgColor="002D62"); GOLD=PatternFill("solid", fgColor="C5A059")
BOLD=Font(bold=True); WHITE=Font(bold=True, color="FFFFFF")

def hdr(ws, row, cols, start=1, fill=NAVY, font=WHITE):
    for j,c in enumerate(cols):
        cell=ws.cell(row=row, column=start+j, value=c); cell.fill=fill; cell.font=font

TICKERS=["AAPL","MSFT","NVDA","GOOGL","AMZN","META","TSLA","AVGO","LLY","JPM",
         "V","UNH","XOM","JNJ","PG","MA","HD","COST","ABBV","MRK"]
COS={"AAPL":"Apple Inc","MSFT":"Microsoft Corp","NVDA":"NVIDIA Corp","GOOGL":"Alphabet Inc",
     "AMZN":"Amazon.com Inc","META":"Meta Platforms","TSLA":"Tesla Inc","AVGO":"Broadcom Inc",
     "LLY":"Eli Lilly","JPM":"JPMorgan Chase","V":"Visa Inc","UNH":"UnitedHealth","XOM":"Exxon Mobil",
     "JNJ":"Johnson & Johnson","PG":"Procter & Gamble","MA":"Mastercard","HD":"Home Depot",
     "COST":"Costco","ABBV":"AbbVie","MRK":"Merck"}

# ============================================================================
# DEMO 1 — "Portfolio Dashboard": KPI cards + holdings + transactions (the
#   showcase: multiple tables per sheet, side by side. A flat parser mangles this.)
# ============================================================================
def demo1():
    wb=Workbook(); ws=wb.active; ws.title="Portfolio"
    ws["B2"]="Q4 2025 Portfolio Snapshot"; ws["B2"].font=Font(bold=True,size=14)
    # KPI strip (labeled 2-row block so it surfaces)
    hdr(ws,4,["Total Value","Total Invested","Total P&L","Return %"],start=2,fill=GOLD,font=BOLD)
    tv=random.randint(800000,1200000); inv=int(tv*random.uniform(0.7,0.9)); pnl=tv-inv
    for j,v in enumerate([tv,inv,pnl,round(pnl/inv,4)]): ws.cell(row=5,column=2+j,value=v)
    # Holdings table (left)
    hdr(ws,8,["Ticker","Company","Shares","Avg Cost","Price","Mkt Value","Weight"],start=2)
    picks=random.sample(TICKERS,8); tot=0; rows=[]
    for t in picks:
        sh=random.randint(10,200); ac=random.randint(50,400); px=round(ac*random.uniform(0.7,2.2),2)
        mv=round(sh*px,2); tot+=mv; rows.append([t,COS[t],sh,ac,px,mv])
    for i,r in enumerate(rows):
        w=round(r[5]/tot,4)
        for j,v in enumerate(r+[w]): ws.cell(row=9+i,column=2+j,value=v)
    # Transactions log (right, side-by-side)
    hdr(ws,8,["Date","Action","Ticker","Qty","Amount"],start=10)
    acts=["BUY","SELL","DIV","BUY","SELL","BUY","DIV"]
    for i,a in enumerate(acts):
        t=random.choice(picks); amt=round(random.uniform(-5000,8000),2)
        for j,v in enumerate([f"2025-11-{i+3:02d}",a,t,random.randint(1,50),amt]):
            ws.cell(row=9+i,column=10+j,value=v)
    wb.save(f"{D}/1_Portfolio_Dashboard.xlsx")
    return ("1_Portfolio_Dashboard.xlsx",
            ["What's my total portfolio value and return?",
             "Which holding has the largest weight?",
             "List all my SELL transactions."])

# ============================================================================
# DEMO 2 — "Financial Statements": 3 sheets, real hierarchy w/ subtotals.
#   Showcases the STATEMENT GUARD (won't wrongly sum a balance sheet).
# ============================================================================
def demo2():
    wb=Workbook(); wb.remove(wb.active)
    def stmt(name, rows):
        ws=wb.create_sheet(name); hdr(ws,1,["Line Item","2023","2024","2025"])
        for i,r in enumerate(rows):
            for j,v in enumerate(r): ws.cell(row=2+i,column=1+j,value=v)
    stmt("Income Statement",[
        ["Revenue",120000,145000,172000],["COGS",-48000,-56000,-64000],
        ["Gross Profit",72000,89000,108000],["Operating Expenses",-35000,-41000,-48000],
        ["Operating Income",37000,48000,60000],["Taxes",-7400,-9600,-12000],
        ["Net Income",29600,38400,48000]])
    stmt("Balance Sheet",[
        ["Assets",None,None,None],["Cash",50000,62000,80000],["Receivables",18000,21000,24000],
        ["Total Current Assets",68000,83000,104000],["PP&E",120000,140000,155000],
        ["Total Assets",188000,223000,259000],["Liabilities",None,None,None],
        ["Payables",22000,26000,30000],["Long-Term Debt",40000,38000,35000],
        ["Total Liabilities",62000,64000,65000],["Total Equity",126000,159000,194000]])
    stmt("Cash Flow",[
        ["Operating",42000,51000,63000],["Investing",-30000,-25000,-28000],
        ["Financing",-8000,-10000,-12000],["Net Change in Cash",4000,16000,23000]])
    wb.save(f"{D}/2_Financial_Statements.xlsx")
    return ("2_Financial_Statements.xlsx",
            ["What was Net Income in 2025?",
             "What were Total Assets each year?",
             "Compute the gross margin for 2024."])

# ============================================================================
# DEMO 3 — "Sales by Region": clean flat table, big enough to show aggregation.
#   Showcases exact stats over ALL rows (avg/median/min/max), not just a sample.
# ============================================================================
def demo3():
    wb=Workbook(); ws=wb.active; ws.title="Sales"
    hdr(ws,1,["Order ID","Date","Region","Product","Rep","Units","Unit Price","Revenue"])
    regions=["North","South","East","West","Central"]; prods=["Alpha","Beta","Gamma","Delta"]
    reps=["Lee","Patel","Garcia","Kim","Novak","Osei"]
    for i in range(120):
        u=random.randint(1,80); pr=round(random.uniform(20,500),2)
        row=[1000+i,f"2025-{random.randint(1,12):02d}-{random.randint(1,28):02d}",
             random.choice(regions),random.choice(prods),random.choice(reps),
             u,pr,round(u*pr,2)]
        for j,v in enumerate(row): ws.cell(row=2+i,column=1+j,value=v)
    wb.save(f"{D}/3_Sales_by_Region.xlsx")
    return ("3_Sales_by_Region.xlsx",
            ["What's the total revenue across all 120 orders?",
             "Which region has the highest average order value?",
             "Who is the top sales rep by revenue?"])

# ============================================================================
# DEMO 4 — "Multi-Fund Comparison": several fund holding sheets (like your real
#   files) — showcases ROUTING (ask about one fund among many).
# ============================================================================
def demo4():
    wb=Workbook(); wb.remove(wb.active)
    funds={"Tech Growth Fund":TICKERS[:6],"Healthcare Fund":["LLY","UNH","JNJ","ABBV","MRK"],
           "Dividend Fund":["PG","JPM","XOM","V","MA","HD"],"Balanced Fund":random.sample(TICKERS,7)}
    for name,ts in funds.items():
        ws=wb.create_sheet(name[:31]); hdr(ws,1,["Ticker","Company","1Y Return","Weight"])
        tot=0; rows=[]
        for t in ts:
            r=round(random.uniform(-0.25,0.9),4); w=random.uniform(0.05,0.25); tot+=w
            rows.append([t,COS[t],r,w])
        for i,r in enumerate(rows):
            r[3]=round(r[3]/tot,4)
            for j,v in enumerate(r): ws.cell(row=2+i,column=1+j,value=v)
    wb.save(f"{D}/4_Multi_Fund_Comparison.xlsx")
    return ("4_Multi_Fund_Comparison.xlsx",
            ["What's in the Healthcare Fund?",
             "Which fund has the best 1-year return on its top holding?",
             "Compare the top holding of Tech Growth Fund and Dividend Fund."])

demos=[demo1(),demo2(),demo3(),demo4()]
print("Generated demo workbooks + suggested questions:\n")
for fn,qs in demos:
    print(f"### {fn}")
    for q in qs: print(f"    - {q}")
    print()
