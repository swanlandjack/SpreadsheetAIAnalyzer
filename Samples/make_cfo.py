"""
Realistic 10-K style financial workbook for a fictional public company
('Northwind Technologies Inc.', ticker NWND). Models what a CFO / annual report
actually contains: Income Statement, Balance Sheet, Cash Flow, Statement of
Equity, plus a Payroll/Compensation schedule and Segment breakdown.

Numbers are internally consistent (net income flows to retained earnings;
assets = liabilities + equity) so it stresses cross-statement reasoning too.
Figures in USD thousands unless noted.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import random
random.seed(2026)

NAVY = PatternFill("solid", fgColor="1F3864")
LBLUE = PatternFill("solid", fgColor="D6E4F0")
GREY = PatternFill("solid", fgColor="F2F2F2")
WHITEB = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
ITAL = Font(italic=True, size=9, color="666666")
thin = Side(style="thin", color="BBBBBB")
TOPBORDER = Border(top=thin)

YEARS = [2023, 2024, 2025]

def sheet(wb, name):
    ws = wb.create_sheet(name[:31])
    ws.sheet_view.showGridLines = False
    return ws

def title_block(ws, subtitle):
    ws["A1"] = "Northwind Technologies Inc. (NASDAQ: NWND)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, color="555555")
    ws["A3"] = "(in thousands of USD, except per-share data)"
    ws["A3"].font = ITAL

def hdr_row(ws, r, extra_left="Line Item"):
    ws.cell(r, 1, extra_left).font = WHITEB
    ws.cell(r, 1).fill = NAVY
    for j, y in enumerate(YEARS):
        c = ws.cell(r, 2+j, y)
        c.font = WHITEB; c.fill = NAVY; c.alignment = Alignment(horizontal="right")

def line(ws, r, label, vals, bold=False, indent=0, border=False, section=False):
    lc = ws.cell(r, 1, ("    "*indent) + label)
    if bold: lc.font = BOLD
    if section:
        lc.font = BOLD; lc.fill = LBLUE
        for j in range(len(YEARS)): ws.cell(r, 2+j).fill = LBLUE
        return
    for j, v in enumerate(vals):
        c = ws.cell(r, 2+j, v)
        c.number_format = '#,##0;(#,##0)'
        if bold: c.font = BOLD
        if border: c.border = TOPBORDER

wb = Workbook(); wb.remove(wb.active)

# ============================================================================
# INCOME STATEMENT
# ============================================================================
ws = sheet(wb, "Income Statement"); title_block(ws, "Consolidated Statements of Operations")
r = 5; hdr_row(ws, r); r += 1
rev = [842000, 968000, 1105000]
cogs = [-378900, -426000, -475000]
gp = [rev[i]+cogs[i] for i in range(3)]
rnd = [-98000, -112000, -128000]; sga = [-142000, -158000, -176000]; mkt = [-64000, -71000, -79000]
opex_total = [rnd[i]+sga[i]+mkt[i] for i in range(3)]
op_income = [gp[i]+opex_total[i] for i in range(3)]
interest = [-12000, -11000, -9500]; other = [3000, 4200, 5100]
pretax = [op_income[i]+interest[i]+other[i] for i in range(3)]
tax = [round(-pretax[i]*0.21) for i in range(3)]
net = [pretax[i]+tax[i] for i in range(3)]
shares = [210000, 212000, 214000]
eps = [round(net[i]/shares[i], 2) for i in range(3)]

line(ws, r, "Revenue", rev, bold=True); r+=1
line(ws, r, "Cost of Revenue", cogs, indent=1); r+=1
line(ws, r, "Gross Profit", gp, bold=True, border=True); r+=1
r+=0
line(ws, r, "Operating Expenses", None, section=True); r+=1
line(ws, r, "Research & Development", rnd, indent=1); r+=1
line(ws, r, "Selling, General & Administrative", sga, indent=1); r+=1
line(ws, r, "Marketing", mkt, indent=1); r+=1
line(ws, r, "Total Operating Expenses", opex_total, bold=True, border=True); r+=1
line(ws, r, "Operating Income", op_income, bold=True); r+=1
line(ws, r, "Interest Expense", interest, indent=1); r+=1
line(ws, r, "Other Income, net", other, indent=1); r+=1
line(ws, r, "Income Before Taxes", pretax, bold=True, border=True); r+=1
line(ws, r, "Provision for Income Taxes", tax, indent=1); r+=1
line(ws, r, "Net Income", net, bold=True, border=True); r+=2
line(ws, r, "Diluted EPS ($)", eps); r+=1
line(ws, r, "Diluted Shares Outstanding", shares); r+=1

# ============================================================================
# BALANCE SHEET
# ============================================================================
ws = sheet(wb, "Balance Sheet"); title_block(ws, "Consolidated Balance Sheets")
r = 5; hdr_row(ws, r); r += 1
cash=[156000,198000,241000]; sti=[89000,102000,120000]; ar=[112000,131000,148000]
inv=[64000,71000,78000]; prepaid=[18000,21000,23000]
tca=[cash[i]+sti[i]+ar[i]+inv[i]+prepaid[i] for i in range(3)]
ppe=[298000,342000,388000]; goodwill=[145000,145000,168000]; intang=[62000,54000,71000]
lti=[38000,44000,52000]; other_a=[21000,24000,27000]
tnca=[ppe[i]+goodwill[i]+intang[i]+lti[i]+other_a[i] for i in range(3)]
ta=[tca[i]+tnca[i] for i in range(3)]
ap=[68000,79000,88000]; accr=[42000,48000,54000]; defrev=[54000,63000,74000]; std=[25000,20000,15000]
tcl=[ap[i]+accr[i]+defrev[i]+std[i] for i in range(3)]
ltd=[180000,165000,150000]; dtl=[34000,38000,41000]; other_l=[19000,22000,25000]
tncl=[ltd[i]+dtl[i]+other_l[i] for i in range(3)]
tl=[tcl[i]+tncl[i] for i in range(3)]
cs=[2100,2120,2140]; apic=[312000,318000,326000]
# retained earnings rolls forward with net income (consistency)
tsy=[-45000,-45000,-52000]
# Retained Earnings is the PLUG so the sheet balances: TA = TL + TEQ, and
# TEQ = CS + APIC + RE + Treasury  ->  RE = (TA - TL) - CS - APIC - Treasury
re=[(ta[i]-tl[i]) - cs[i] - apic[i] - tsy[i] for i in range(3)]
teq=[cs[i]+apic[i]+re[i]+tsy[i] for i in range(3)]
tle=[tl[i]+teq[i] for i in range(3)]

line(ws, r, "ASSETS", None, section=True); r+=1
line(ws, r, "Current Assets", None, section=True); r+=1
line(ws, r, "Cash & Cash Equivalents", cash, indent=1); r+=1
line(ws, r, "Short-Term Investments", sti, indent=1); r+=1
line(ws, r, "Accounts Receivable, net", ar, indent=1); r+=1
line(ws, r, "Inventory", inv, indent=1); r+=1
line(ws, r, "Prepaid Expenses", prepaid, indent=1); r+=1
line(ws, r, "Total Current Assets", tca, bold=True, border=True); r+=1
line(ws, r, "Property, Plant & Equipment, net", ppe, indent=1); r+=1
line(ws, r, "Goodwill", goodwill, indent=1); r+=1
line(ws, r, "Intangible Assets, net", intang, indent=1); r+=1
line(ws, r, "Long-Term Investments", lti, indent=1); r+=1
line(ws, r, "Other Non-Current Assets", other_a, indent=1); r+=1
line(ws, r, "Total Non-Current Assets", tnca, bold=True, border=True); r+=1
line(ws, r, "TOTAL ASSETS", ta, bold=True, border=True); r+=2
line(ws, r, "LIABILITIES & EQUITY", None, section=True); r+=1
line(ws, r, "Current Liabilities", None, section=True); r+=1
line(ws, r, "Accounts Payable", ap, indent=1); r+=1
line(ws, r, "Accrued Liabilities", accr, indent=1); r+=1
line(ws, r, "Deferred Revenue", defrev, indent=1); r+=1
line(ws, r, "Short-Term Debt", std, indent=1); r+=1
line(ws, r, "Total Current Liabilities", tcl, bold=True, border=True); r+=1
line(ws, r, "Long-Term Debt", ltd, indent=1); r+=1
line(ws, r, "Deferred Tax Liabilities", dtl, indent=1); r+=1
line(ws, r, "Other Non-Current Liabilities", other_l, indent=1); r+=1
line(ws, r, "Total Non-Current Liabilities", tncl, bold=True, border=True); r+=1
line(ws, r, "Total Liabilities", tl, bold=True, border=True); r+=1
line(ws, r, "Shareholders' Equity", None, section=True); r+=1
line(ws, r, "Common Stock", cs, indent=1); r+=1
line(ws, r, "Additional Paid-In Capital", apic, indent=1); r+=1
line(ws, r, "Retained Earnings", re, indent=1); r+=1
line(ws, r, "Treasury Stock", tsy, indent=1); r+=1
line(ws, r, "Total Shareholders' Equity", teq, bold=True, border=True); r+=1
line(ws, r, "TOTAL LIABILITIES & EQUITY", tle, bold=True, border=True); r+=1

# ============================================================================
# CASH FLOW
# ============================================================================
ws = sheet(wb, "Cash Flow"); title_block(ws, "Consolidated Statements of Cash Flows")
r=5; hdr_row(ws, r); r+=1
dep=[48000,54000,61000]; sbc=[32000,38000,44000]; wc=[-18000,-22000,-15000]
cfo=[net[i]+dep[i]+sbc[i]+wc[i] for i in range(3)]
capex=[-72000,-88000,-94000]; acq=[0,0,-45000]; invsec=[-14000,-16000,-20000]
cfi=[capex[i]+acq[i]+invsec[i] for i in range(3)]
div=[round(-net[i]*0.15) for i in range(3)]; buyback=[0,0,-7000]; debt=[-15000,-15000,-15000]
cff=[div[i]+buyback[i]+debt[i] for i in range(3)]
netchg=[cfo[i]+cfi[i]+cff[i] for i in range(3)]

line(ws, r, "Operating Activities", None, section=True); r+=1
line(ws, r, "Net Income", net, indent=1); r+=1
line(ws, r, "Depreciation & Amortization", dep, indent=1); r+=1
line(ws, r, "Stock-Based Compensation", sbc, indent=1); r+=1
line(ws, r, "Changes in Working Capital", wc, indent=1); r+=1
line(ws, r, "Net Cash from Operating Activities", cfo, bold=True, border=True); r+=1
line(ws, r, "Investing Activities", None, section=True); r+=1
line(ws, r, "Capital Expenditures", capex, indent=1); r+=1
line(ws, r, "Acquisitions, net of cash", acq, indent=1); r+=1
line(ws, r, "Purchases of Investments", invsec, indent=1); r+=1
line(ws, r, "Net Cash from Investing Activities", cfi, bold=True, border=True); r+=1
line(ws, r, "Financing Activities", None, section=True); r+=1
line(ws, r, "Dividends Paid", div, indent=1); r+=1
line(ws, r, "Share Repurchases", buyback, indent=1); r+=1
line(ws, r, "Repayment of Debt", debt, indent=1); r+=1
line(ws, r, "Net Cash from Financing Activities", cff, bold=True, border=True); r+=1
line(ws, r, "Net Change in Cash", netchg, bold=True, border=True); r+=1

# ============================================================================
# PAYROLL / COMPENSATION SCHEDULE (flat table — different shape)
# ============================================================================
ws = sheet(wb, "Payroll Summary"); title_block(ws, "Compensation Schedule — FY2025 (Named Executive Officers & Departments)")
r=5
cols=["Employee / Group","Title / Dept","Base Salary","Bonus","Stock Awards","Benefits","Total Comp","Headcount"]
for j,c in enumerate(cols):
    cell=ws.cell(r,1+j,c); cell.font=WHITEB; cell.fill=NAVY
r+=1
neos=[
 ("A. Rivera","Chief Executive Officer",1200,2400,8500,180,1),
 ("M. Chen","Chief Financial Officer",750,1100,3200,120,1),
 ("K. Okafor","Chief Technology Officer",820,1250,3800,130,1),
 ("S. Nakamura","Chief Operating Officer",700,980,2900,115,1),
 ("D. Volkov","General Counsel",620,760,2100,105,1),
]
depts=[
 ("Engineering","Dept — R&D",142000,18000,54000,28000,640),
 ("Sales","Dept — GTM",98000,42000,12000,19000,410),
 ("Customer Success","Dept — CS",56000,8000,4000,11000,280),
 ("G&A","Dept — Corporate",48000,6000,9000,9500,190),
]
for e in neos:
    for j,v in enumerate(e): ws.cell(r,1+j,v)
    ws.cell(r,7,e[2]+e[3]+e[4]+e[5])  # total comp
    r+=1
for d in depts:
    for j,v in enumerate(d): ws.cell(r,1+j,v)
    ws.cell(r,7,d[2]+d[3]+d[4]+d[5])
    r+=1
# footer note row (the kind that trips naive parsers)
ws.cell(r,1,"Note: NEO figures in $ thousands; department figures aggregate all staff. Total company headcount: 1,525.").font=ITAL

# ============================================================================
# SEGMENT DATA (flat, year columns but a REVENUE table not a statement)
# ============================================================================
ws = sheet(wb, "Segments"); title_block(ws, "Revenue by Segment & Geography")
r=5
ws.cell(r,1,"Segment").font=WHITEB; ws.cell(r,1).fill=NAVY
for j,y in enumerate(YEARS):
    c=ws.cell(r,2+j,y); c.font=WHITEB; c.fill=NAVY
r+=1
segs=[("Cloud Platform",[402000,498000,592000]),
      ("Enterprise Software",[268000,289000,312000]),
      ("Professional Services",[112000,121000,131000]),
      ("Hardware",[60000,60000,70000])]
for name,vals in segs:
    ws.cell(r,1,name)
    for j,v in enumerate(vals): ws.cell(r,2+j,v).number_format='#,##0'
    r+=1
ws.cell(r,1,"Total Revenue").font=BOLD
for j in range(3):
    tot=sum(s[1][j] for s in segs)
    c=ws.cell(r,2+j,tot); c.font=BOLD; c.border=TOPBORDER; c.number_format='#,##0'

wb.save("Northwind_10K_Financials.xlsx")
print("saved Northwind_10K_Financials.xlsx")
print("sheets:", wb.sheetnames)
print(f"consistency check: 2025 Assets={ta[2]:,}  L+E={tle[2]:,}  match={ta[2]==tle[2]}")
print(f"2025 Net Income={net[2]:,}  flows to Cash Flow op activities line")
